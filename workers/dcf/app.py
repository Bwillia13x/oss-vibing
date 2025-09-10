from fastapi import FastAPI
from pydantic import BaseModel
from typing import Any, Dict, Optional, List, Tuple, cast
from datetime import datetime, timezone
import os

try:
    import openpyxl  # type: ignore
    HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    openpyxl = None  # type: ignore
    HAS_OPENPYXL = False

app = FastAPI(title="Mad Lab DCF Worker")


class Scenarios(BaseModel):
    base: Dict[str, Any]
    bull: Optional[Dict[str, Any]] = None
    bear: Optional[Dict[str, Any]] = None


class DcfRequest(BaseModel):
    ticker: str
    horizon: int
    wacc: Optional[float] = None  # percent or decimal (10 or 0.10)
    scenarios: Optional[Scenarios] = None


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


def _pct(x: Optional[float], default: float) -> float:
    if x is None:
        return default
    # allow 10 instead of 0.10
    return x / 100.0 if x > 1 else x


def _scenario_defaults(overrides: Optional[Dict[str, Any]], horizon: int, r: float) -> Dict[str, Any]:
    o = overrides or {}
    # Interpret flexible inputs; apply sensible defaults
    return {
        "revenue0": float(o.get("revenue0", 1000.0)),
        "growth": _pct(cast(Optional[float], o.get("growth")), 0.10),  # starting growth
        "growth_decay": _pct(cast(Optional[float], o.get("growth_decay")), 0.02),  # converge toward terminal_g
        "ebit_margin": _pct(cast(Optional[float], o.get("ebit_margin")), 0.20),
        "tax_rate": _pct(cast(Optional[float], o.get("tax_rate")), 0.21),
        "reinvestment_rate": _pct(cast(Optional[float], o.get("reinvestment_rate")), 0.35),
        "terminal_g": _pct(cast(Optional[float], o.get("terminal_g")), 0.02),
        "wacc": _pct(cast(Optional[float], o.get("wacc")), r),
        "shares_out": float(o.get("shares_out", 0.0)),
        "net_debt": float(o.get("net_debt", 0.0)),
        "horizon": int(o.get("horizon", horizon)),
    }


def _project_cash_flows(s: Dict[str, Any]) -> Tuple[List[Dict[str, float]], float, float]:
    N: int = int(s["horizon"])
    rev = s["revenue0"]
    g = float(s["growth"])  # starting growth
    g_terminal = float(s["terminal_g"])  # converge towards this by year N
    g_decay = float(s["growth_decay"])  # yearly reduction in growth until reaching terminal
    margin = float(s["ebit_margin"])
    tax = float(s["tax_rate"])
    reinv = float(s["reinvestment_rate"])
    r = float(s["wacc"])

    flows: List[Dict[str, float]] = []
    pv_sum = 0.0

    # helper to decay growth toward terminal
    def next_growth(curr: float) -> float:
        if curr > g_terminal:
            return max(g_terminal, curr - g_decay)
        else:
            return min(g_terminal, curr + g_decay)

    for t in range(1, N + 1):
        rev = rev * (1.0 + g)
        ebit = rev * margin
        nopat = ebit * (1.0 - tax)
        reinvest = max(0.0, nopat * reinv)
        fcf = nopat - reinvest
        df = 1.0 / ((1.0 + r) ** t)
        pv_fcf = fcf * df
        pv_sum += pv_fcf
        flows.append(
            {
                "year": float(t),
                "revenue": rev,
                "ebit": ebit,
                "nopat": nopat,
                "reinvestment": reinvest,
                "fcf": fcf,
                "discountFactor": df,
                "pv_fcf": pv_fcf,
                "cum_pv": pv_sum,
            }
        )
        g = next_growth(g)

    # Terminal value using Gordon growth on next year's FCF
    last = flows[-1]
    # Estimate next-year FCF: grow NOPAT, apply same reinvestment policy
    nopat_next = last["nopat"] * (1.0 + g_terminal)
    fcf_next = nopat_next * (1.0 - reinv)
    tv = fcf_next / max(1e-6, (r - g_terminal))
    pv_tv = tv / ((1.0 + r) ** N)
    return flows, tv, pv_tv


def _write_workbook(base_dir: str, assumptions: Dict[str, Any], pv_by_scenario: Dict[str, float], sensitivity: List[Dict[str, float]], flows_by_scenario: Dict[str, List[Dict[str, float]]]) -> Optional[str]:
    if not HAS_OPENPYXL:
        return None
    wb = cast(Any, openpyxl).Workbook()
    # Assumptions
    ws1 = wb.active
    ws1.title = "Assumptions"
    ws1.append(["key", "value"])
    for k, v in assumptions.items():
        ws1.append([k, v])
    # PV by scenario
    ws2 = wb.create_sheet(title="PV")
    ws2.append(["scenario", "enterprise_value"])
    for k, v in pv_by_scenario.items():
        ws2.append([k, v])
    # Sensitivity (long form)
    ws3 = wb.create_sheet(title="Sensitivity")
    ws3.append(["wacc", "terminalG", "value"])
    for row in sensitivity:
        ws3.append([row["wacc"], row["terminalG"], row["value"]])
    # Cash flows per scenario
    for name, rows in flows_by_scenario.items():
        ws = wb.create_sheet(title=f"CashFlows-{name}")
        ws.append([
            "year",
            "revenue",
            "ebit",
            "nopat",
            "reinvestment",
            "fcf",
            "discountFactor",
            "pv_fcf",
            "cum_pv",
        ])
        for r in rows:
            ws.append([
                r["year"],
                r["revenue"],
                r["ebit"],
                r["nopat"],
                r["reinvestment"],
                r["fcf"],
                r["discountFactor"],
                r["pv_fcf"],
                r["cum_pv"],
            ])
    xlsx_path = os.path.join(base_dir, "dcf.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


@app.post("/dcf")
def dcf(req: DcfRequest) -> Dict[str, Any]:
    # Normalize r (WACC)
    r = _pct(req.wacc, 0.10)
    N = int(max(1, req.horizon))
    sc = req.scenarios or Scenarios(base={})

    # Build per-scenario inputs
    s_base = _scenario_defaults(sc.base, N, r)
    s_bull = _scenario_defaults(sc.bull or {}, N, r) if sc.bull is not None else None
    s_bear = _scenario_defaults(sc.bear or {}, N, r) if sc.bear is not None else None

    flows_by: Dict[str, List[Dict[str, float]]] = {}
    pv_by_scenario: Dict[str, float] = {}

    # Compute base
    base_flows, _base_tv, base_pv_tv = _project_cash_flows(s_base)
    ev_base = base_flows[-1]["cum_pv"] + base_pv_tv
    flows_by["base"] = base_flows
    pv_by_scenario["base"] = round(ev_base, 2)

    # Optional bull
    if s_bull is not None:
        bull_flows, _, bull_pv_tv = _project_cash_flows(s_bull)
        ev_bull = bull_flows[-1]["cum_pv"] + bull_pv_tv
        flows_by["bull"] = bull_flows
        pv_by_scenario["bull"] = round(ev_bull, 2)
    else:
        pv_by_scenario["bull"] = round(ev_base * 1.2, 2)

    # Optional bear
    if s_bear is not None:
        bear_flows, _, bear_pv_tv = _project_cash_flows(s_bear)
        ev_bear = bear_flows[-1]["cum_pv"] + bear_pv_tv
        flows_by["bear"] = bear_flows
        pv_by_scenario["bear"] = round(ev_bear, 2)
    else:
        pv_by_scenario["bear"] = round(ev_base * 0.8, 2)

    # Sensitivity around base terminal year based on base scenario
    last_base = flows_by["base"][-1]
    base_g = float(s_base["terminal_g"])
    reinv = float(s_base["reinvestment_rate"])
    nopat_next = last_base["nopat"] * (1.0 + base_g)
    fcf_next = nopat_next * (1.0 - reinv)
    sensitivity: List[Dict[str, float]] = []
    for w in range(6, 13):  # 6..12
        for g_i in [0.0, 0.5, 1.0, 1.5, 2.0, 2.5]:
            g = g_i / 100.0
            rr = w / 100.0
            tv = fcf_next / max(1e-6, (rr - g))
            pv_tv = tv / ((1.0 + rr) ** N)
            value = last_base["cum_pv"] + pv_tv
            sensitivity.append({"wacc": float(w), "terminalG": float(g_i), "value": round(float(value), 2)})

    assumptions = {
        "ticker": req.ticker,
        "horizon": N,
        "wacc": round(r * 100.0, 2),
        "notes": "Fundamentals-driven DCF using simple revenue->EBIT->NOPAT and reinvestment policy.",
    }

    files: List[str] = []
    try:
        base_dir = f"/tmp/madlab-dcf-{int(datetime.now(timezone.utc).timestamp()*1000)}"
        os.makedirs(base_dir, exist_ok=True)
        xlsx = _write_workbook(base_dir, assumptions, pv_by_scenario, sensitivity, flows_by)
        if xlsx:
            files.append(xlsx)
    except Exception:
        # best-effort artifacts
        pass

    return {
        "assumptions": assumptions,
        "pvByScenario": pv_by_scenario,
        "sensitivity": sensitivity,
        "files": files,
        "explain": "DCF computed with fundamentals-based cash flow projection and Gordon terminal value.",
    }
